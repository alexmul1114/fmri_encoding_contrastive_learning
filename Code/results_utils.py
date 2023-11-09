# Imports

import os
os.environ["OMP_NUM_THREADS"] = '1'

import numpy as np
from pathlib import Path
from PIL import Image

import torch
import torch.nn as nn
from torch.utils.data import DataLoader, Dataset, random_split
from torchvision.models.feature_extraction import create_feature_extractor, get_graph_node_names
from torchvision import transforms
from torchvision.models import AlexNet_Weights
from sklearn.decomposition import IncrementalPCA
from sklearn.linear_model import LinearRegression
from scipy.stats import pearsonr as corr
import joblib
import time
import itertools

import torchvision
from torchvision.models.feature_extraction import create_feature_extractor

import torchvision.models as models

from torchmetrics.functional import pairwise_cosine_similarity
from scipy.stats import pearsonr as corr
from scipy.stats import ttest_rel
from sklearn.linear_model import LinearRegression, LogisticRegression
from sklearn.svm import LinearSVC

from sklearn.preprocessing import StandardScaler
from sklearn.metrics import accuracy_score

from sklearn.decomposition import PCA

from tqdm import tqdm

import torchextractor as tx
from sklearn.linear_model import Ridge
from sklearn.svm import LinearSVC

from sklearn.manifold import TSNE

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import xlwings

import matplotlib.pyplot as plt
import seaborn as sns

# Local imports
from utils import get_dataloaders, fit_pca, extract_pca_features, get_fmri_from_dataloader, get_dataloaders_with_img_paths, get_dataloaders_cv
#from algonauts_models import CLR_model
from models import CLR_model, fmri_reg



# Function to print results for control linear encoding, CL alex fine-tuned feature 
# encoding with regularization, and neural network regression alex fined-tuned feature
# encoding with regularization, all for subject-specific models
def get_results_single_subj(project_dir, device, subj_num, hemisphere, roi):
    
    hemisphere_abbr = 'l' if hemisphere=='left' else 'r'
    print(roi, hemisphere)
    # Get dataloaders
    train_dataloader, test_dataloader, train_size, test_size, num_voxels = get_dataloaders(project_dir, 
                                                                device, subj_num, hemisphere, roi, 1024, shuffle=False)
    if (num_voxels==0):
        print("Empty ROI")
        return -1,-1,-1,-1,-1,-1,-1
    elif (num_voxels<20):
        print("Too few voxels")
        return -1,-1,-1,-1,-1,-1,-1
    
    # Get training and testing fmri as numpy arrays
    train_fmri = get_fmri_from_dataloader(train_dataloader, train_size, num_voxels)
    test_fmri = get_fmri_from_dataloader(test_dataloader, test_size, num_voxels)
    
    
    # Get control linear encoding model predictions:
    print("Getting untuned predictions...")
    # Load best alexnet layer for control linear encoding model, create feature extractor
    #best_alex_out_layer_path = project_dir + r"/best_alex_out_layers/Subj" + str(subj_num) + r"/subj" + str(subj_num) + "_" + hemisphere_abbr + "h_best_alex_layer_dict.joblib"
    best_alex_out_layer_path = project_dir + r"/best_alex_out_layers/best_alex_layer_dict.joblib"
    best_alex_out_layer_dict = joblib.load(best_alex_out_layer_path)
    alex_out_layer_dims = {"features.2":46656, "features.5":32448, "features.7":64896, "features.9":43264,
                           "features.12":9216, "classifier.2":4096, "classifier.5":4096, "classifier.6":1000}
    roi_out_name = hemisphere_abbr + "h_" + roi
    best_alex_layer, best_untuned_alpha = best_alex_out_layer_dict[roi_out_name]
    alex_out_size = alex_out_layer_dims[best_alex_layer]
    
    alex = torch.hub.load('pytorch/vision:v0.10.0', 'alexnet', weights=AlexNet_Weights.IMAGENET1K_V1)
    alex.to(device) # send the model to the chosen device 
    alex.eval() # set the model to evaluation mode
    feature_extractor = create_feature_extractor(alex, return_nodes=[best_alex_layer])

    del alex
    
    #Fit PCA using feature extractor
    pca = fit_pca(feature_extractor, train_dataloader, train_size, best_alex_layer, alex_out_size)
    
    # Get training and testing image pca features
    
    train_pca_features = extract_pca_features(feature_extractor, train_dataloader, pca, best_alex_layer, train_size)
    test_pca_features = extract_pca_features(feature_extractor, test_dataloader, pca, best_alex_layer, test_size)
    
    # Fit control linear encoding model, get test predictions
    control_linear_model = Ridge(alpha=best_untuned_alpha).fit(train_pca_features, train_fmri)
    
    ctrl_preds = control_linear_model.predict(test_pca_features)

    del feature_extractor, train_pca_features, test_pca_features, control_linear_model
    

    
    # Get preds using tuned alexnet from CL:
    # Load CL model
    print("Getting CL predictions...")
    cl_model_dir = project_dir + r"/cl_models/Subj" + str(subj_num)
    cl_model_path = cl_model_dir + r"/subj" + str(subj_num) + "_" + hemisphere_abbr + "h_" + roi + "_model_e30.pt" 
    h_dim = int(num_voxels*0.8)
    z_dim = int(num_voxels*0.2)
    cl_model = CLR_model(num_voxels, h_dim, z_dim)
    # Some models seem to be saved differently
    try:
        cl_model.load_state_dict(torch.load(cl_model_path)[0].state_dict())
    except:
        try:
            cl_model.load_state_dict(torch.load(cl_model_path).state_dict())
        except:
            cl_model.load_state_dict(torch.load(cl_model_path))
    cl_model.to(device)
    cl_model.eval()
    cl_layer = "alex." + best_alex_layer
    feature_extractor = tx.Extractor(cl_model, [cl_layer])
    pca = fit_pca(feature_extractor, train_dataloader, train_size, cl_layer, alex_out_size, 
                is_cl_feature_extractor=True, num_voxels=num_voxels)
    train_features = extract_pca_features(feature_extractor, train_dataloader, pca, cl_layer, train_size, 
                    is_cl_feature_extractor=True, num_voxels=num_voxels)
    test_features = extract_pca_features(feature_extractor, test_dataloader, pca, cl_layer, test_size, 
                    is_cl_feature_extractor=True, num_voxels=num_voxels)

    #cl_encoding_model = Ridge(alpha=10000).fit(train_features, train_fmri)
    cl_encoding_model = Ridge(alpha=best_untuned_alpha).fit(train_features, train_fmri)
    cl_preds = cl_encoding_model.predict(test_features)

    del feature_extractor, train_features, test_features, cl_encoding_model
    
    


    print("Getting regression predictions...")
    reg_model_dir = project_dir + r"/baseline_models/nn_reg/Subj" + str(subj_num)
    reg_model_path = reg_model_dir + r"/subj" + str(subj_num) + "_" + hemisphere_abbr + "h_" + roi + "_reg_model_e75.pt" 
    reg_model = fmri_reg(num_voxels)
    try:
        reg_model.load_state_dict(torch.load(reg_model_path)[0].state_dict())
    except:
        try:
            reg_model.load_state_dict(torch.load(reg_model_path).state_dict())
        except:
            reg_model.load_state_dict(torch.load(reg_model_path))
    reg_model.to(device)
    reg_model.eval()
    #reg_model = torch.load(reg_model_path).to(device)

    # Create feature extractor
    feature_extractor = tx.Extractor(reg_model, [cl_layer])
    pca = fit_pca(feature_extractor, train_dataloader, train_size, cl_layer, alex_out_size, 
                is_cl_feature_extractor=False, is_reg_feature_extractor=True, num_voxels=num_voxels)
    train_features = extract_pca_features(feature_extractor, train_dataloader, pca, cl_layer, train_size, 
                    is_cl_feature_extractor=False, is_reg_feature_extractor=True, num_voxels=num_voxels)
    test_features = extract_pca_features(feature_extractor, test_dataloader, pca, cl_layer, test_size, 
                    is_cl_feature_extractor=False, is_reg_feature_extractor=True, num_voxels=num_voxels)

    #reg_encoding_model = Ridge(alpha=100).fit(train_features, train_fmri)
    reg_encoding_model = Ridge(alpha=best_untuned_alpha).fit(train_features, train_fmri)
    reg_preds = reg_encoding_model.predict(test_features)

    del feature_extractor, train_features, test_features, reg_encoding_model
    
    
    
    # Compute mean correlations for all methods
    ctrl_corrs = np.zeros(num_voxels)
    cl_corrs = np.zeros(num_voxels)
    reg_corrs = np.zeros(num_voxels)
    print("Computing correlations...")
    for v in tqdm(range(num_voxels)):
        ctrl_corrs[v] = corr(test_fmri[:, v], ctrl_preds[:, v])[0]
        cl_corrs[v] = corr(test_fmri[:, v], cl_preds[:, v])[0]
        reg_corrs[v] = corr(test_fmri[:, v], reg_preds[:, v])[0]

    ctrl_avg = ctrl_corrs.mean()
    cl_avg = cl_corrs.mean()
    reg_avg = reg_corrs.mean()
    
    # Get percentage of improved voxels
    cl_ctrl_improved_percentage = np.count_nonzero(cl_corrs - ctrl_corrs > 0) / num_voxels
    cl_reg_improved_percentage = np.count_nonzero(cl_corrs - reg_corrs > 0) / num_voxels
    
    cl_vs_ctrl_voxel_differences = cl_corrs - ctrl_corrs
    cl_vs_reg_voxel_differences = cl_corrs - reg_corrs
    
    print(roi + ":")
    print("Average control correlation: " + str(np.round(ctrl_avg,3)))
    print("Average regression correlation: " + str(np.round(reg_avg,3)))
    print("Average CL correlation: " + str(np.round(cl_avg,3)))

    
    print("% of voxels improved versus control: " + str(np.round(cl_ctrl_improved_percentage,3)))
    print("% of voxels improved versus regression: " + str(np.round(cl_reg_improved_percentage,3)))
    
    del train_fmri, test_fmri, train_dataloader, test_dataloader
    return ctrl_avg, reg_avg, cl_avg, cl_ctrl_improved_percentage, cl_reg_improved_percentage, cl_vs_ctrl_voxel_differences, cl_vs_reg_voxel_differences
    

def get_results_single_subj_all_rois(project_dir, device, subj_num, hemisphere, save=True, exclude_rois=None):
    
    all_rois = ["V1v", "V1d", "V2v", "V2d", "V3v", "V3d", "hV4", "EBA", "FBA-1", "FBA-2",
            "mTL-bodies", "OFA", "FFA-1", "FFA-2", "mTL-faces", "aTL-faces", "OPA",
         "PPA", "RSC", "OWFA", "VWFA-1", "VWFA-2", "mfs-words", "mTL-words",
       "early", "midventral", "midlateral", "midparietal", "ventral", "lateral", "parietal"]

    
    hemisphere_abbr = 'l' if hemisphere=='left' else 'r'

    if (save):
        ctrl_results = {}
        reg_results = {}
        cl_results = {}
        improved_percentages_vs_ctrl = {}
        improved_percentages_vs_reg = {}
        voxel_differences_vs_ctrl = {}
        voxel_differences_vs_reg = {}
          
    for roi in all_rois:
        ctrl_avg, reg_avg, cl_avg, cl_ctrl_improved_percentage, cl_reg_improved_percentage, cl_vs_ctrl_voxel_differences, cl_vs_reg_voxel_differences = get_results_single_subj(project_dir, device, subj_num, hemisphere, roi)
        if (save):
            ctrl_results[roi] = ctrl_avg
            reg_results[roi] = reg_avg
            cl_results[roi] = cl_avg
            improved_percentages_vs_ctrl[roi] = cl_ctrl_improved_percentage
            improved_percentages_vs_reg[roi] = cl_reg_improved_percentage
            voxel_differences_vs_ctrl[roi] = cl_vs_ctrl_voxel_differences
            voxel_differences_vs_reg[roi] = cl_vs_reg_voxel_differences

    # Save results
    if (save):
        
        save_folder = project_dir + "/results/Subj" + str(subj_num) 
        
        ctrl_results_save_file = save_folder + "/subj" + str(subj_num) + "_" + hemisphere_abbr + "h_ctrl_results.joblib"
        cl_results_save_file = save_folder + "/subj" + str(subj_num) + "_" + hemisphere_abbr + "h_cl_results.joblib"
        reg_results_save_file = save_folder + "/subj" + str(subj_num) + "_" + hemisphere_abbr + "h_reg_results.joblib"
        ctrl_improved_percentages_save_file = save_folder + "/subj" + str(subj_num) + "_" + hemisphere_abbr + "h_cl_vs_ctrl_improved_percentages.joblib"
        reg_improved_percentages_save_file = save_folder + "/subj" + str(subj_num) + "_" + hemisphere_abbr + "h_cl_vs_reg_improved_percentages.joblib"
        ctrl_improved_voxels_save_file = save_folder + "/subj" + str(subj_num) + "_" + hemisphere_abbr + "h_cl_vs_ctrl_voxel_differences.joblib"
        reg_improved_voxels_save_file = save_folder + "/subj" + str(subj_num) + "_" + hemisphere_abbr + "h_cl_vs_reg_voxel_differences.joblib"
        
        joblib.dump(ctrl_results, ctrl_results_save_file)
        joblib.dump(cl_results, cl_results_save_file)
        joblib.dump(reg_results, reg_results_save_file)
        joblib.dump(improved_percentages_vs_ctrl, ctrl_improved_percentages_save_file)
        joblib.dump(improved_percentages_vs_reg, reg_improved_percentages_save_file)
        joblib.dump(voxel_differences_vs_ctrl, ctrl_improved_voxels_save_file)
        joblib.dump(voxel_differences_vs_reg, reg_improved_voxels_save_file)
        
        
        
# Function to get results for using models cross-subject
def get_results_cross_subj(project_dir, device, hemisphere, roi, save=True, num_subjs=8, training_results=False):
    
    hemisphere_abbr = 'l' if hemisphere=='left' else 'r'
    
    results_mat = np.zeros((num_subjs, num_subjs))
    voxel_improvement_percentages_ctrl_mat = np.zeros((num_subjs, num_subjs))

    # Get best alex out layer and corresponding dimension for this ROI
    best_alex_out_layer_path = project_dir + r"/best_alex_out_layers/best_alex_layer_dict.joblib"
    best_alex_out_layer_dict = joblib.load(best_alex_out_layer_path)
    alex_out_layer_dims = {"features.2":46656, "features.5":32448, "features.7":64896, "features.9":43264,
                        "features.12":9216, "classifier.2":4096, "classifier.5":4096, "classifier.6":1000}
    roi_out_name = hemisphere_abbr + "h_" + roi
    best_alex_layer, best_untuned_alpha = best_alex_out_layer_dict[roi_out_name]
    alex_out_size = alex_out_layer_dims[best_alex_layer]
    cl_layer = "alex." + best_alex_layer

    
    # Get control predictions for this ROI for each subject
    ctrl_preds_all_subjs = []
    for subj_idx in range(num_subjs):
        
        subj_num = subj_idx + 1

        print("Loading fmri data...")
        train_dataloader, test_dataloader, train_size, test_size, num_voxels = get_dataloaders(project_dir, device, subj_num, hemisphere, roi, 1024, shuffle=False)
        train_fmri = get_fmri_from_dataloader(train_dataloader, train_size, num_voxels)
        test_fmri = get_fmri_from_dataloader(test_dataloader, test_size, num_voxels)
        
        print("Getting control untuned predictions...")
        alex = torch.hub.load('pytorch/vision:v0.10.0', 'alexnet', weights=AlexNet_Weights.IMAGENET1K_V1)
        alex.to(device) # send the model to the chosen device 
        alex.eval() # set the model to evaluation mode
        feature_extractor = create_feature_extractor(alex, return_nodes=[best_alex_layer])
        del alex
        
        
        #Fit PCA using feature extractor
        pca = fit_pca(feature_extractor, train_dataloader, train_size, best_alex_layer, alex_out_size)
    
        # Get training and testing image pca features
        train_pca_features = extract_pca_features(feature_extractor, train_dataloader, pca, best_alex_layer, train_size)
        test_pca_features = extract_pca_features(feature_extractor, test_dataloader, pca, best_alex_layer, test_size)
        
        # Fit control linear encoding model, get test predictions
        control_linear_model = Ridge(alpha=best_untuned_alpha).fit(train_pca_features, train_fmri)

        if (training_results):
            ctrl_preds = control_linear_model.predict(train_pca_features)
        else:
            ctrl_preds = control_linear_model.predict(test_pca_features)
        
        ctrl_preds_all_subjs.append(ctrl_preds)
    
        
    
    for model_subj_idx in range(num_subjs):

        model_subj = model_subj_idx + 1

        # Need the number of voxels in the ROI for the model subject
        _, _, _, _, num_voxels_model_subj = get_dataloaders(project_dir, device, model_subj, hemisphere, roi, 1024, shuffle=False)

        cl_model_dir = project_dir + r"/cl_models/Subj" + str(model_subj)
        cl_model_path = cl_model_dir + r"/subj" + str(model_subj) + "_" + hemisphere_abbr + "h_" + roi + "_model_e30.pt" 
        h_dim = int(num_voxels_model_subj*0.8)
        z_dim = int(num_voxels_model_subj*0.2)
        cl_model = CLR_model(num_voxels_model_subj, h_dim, z_dim)
        # Some models seem to be saved differently
        try:
            cl_model.load_state_dict(torch.load(cl_model_path, map_location=torch.device('cpu'))[0].state_dict())
        except:
            try:
                cl_model.load_state_dict(torch.load(cl_model_path, map_location=torch.device('cpu')).state_dict())
            except:
                cl_model.load_state_dict(torch.load(cl_model_path, map_location=torch.device('cpu')))
        cl_model.to(device)
        cl_model.eval()
        feature_extractor = tx.Extractor(cl_model, [cl_layer])
        del cl_model
        

        for target_subj_idx in range(num_subjs):
            
            target_subj = target_subj_idx + 1
            
            # Get dataloaders for target subj
            train_dataloader, test_dataloader, train_size, test_size, num_voxels_target_subj = get_dataloaders(project_dir, device, target_subj, hemisphere, roi, 1024, shuffle=False)
            
            if (num_voxels_target_subj==0):
                results_mat[model_subj_idx, target_subj_idx] = -1
            else:

                # Get training and testing fmri for target subj
                print("Loading fmri data...")
                train_fmri = get_fmri_from_dataloader(train_dataloader, train_size, num_voxels_target_subj)
                test_fmri = get_fmri_from_dataloader(test_dataloader, test_size, num_voxels_target_subj)

                # Get image features for target subj images
                print("Getting image features...")
                pca = fit_pca(feature_extractor, train_dataloader, train_size, cl_layer, alex_out_size, is_cl_feature_extractor=True, 
                              make_dummy_fmri_data=True, num_voxels=num_voxels_model_subj)
                train_features = extract_pca_features(feature_extractor, train_dataloader, pca, cl_layer, train_size, 
                                is_cl_feature_extractor=True, make_dummy_fmri_data=True, num_voxels=num_voxels_model_subj)
                test_features = extract_pca_features(feature_extractor, test_dataloader, pca, cl_layer, test_size,  
                                is_cl_feature_extractor=True, make_dummy_fmri_data=True, num_voxels=num_voxels_model_subj)

            
                # Fit encoding model and use it to predict test fmri
                print("Fitting linear encoding model...")
                encoding_model = Ridge(alpha=best_untuned_alpha).fit(train_features, train_fmri)
                
                if (training_results):
                    cs_preds = encoding_model.predict(train_features)
                else:
                    cs_preds = encoding_model.predict(test_features)

                # Evaluate mean correlation accuracy
                ctrl_corrs = np.zeros(num_voxels_target_subj)
                cs_corrs = np.zeros(num_voxels_target_subj)
                
                ctrl_preds = ctrl_preds_all_subjs[target_subj_idx]
                print("Computing correlations...")
                for v in tqdm(range(num_voxels_target_subj)):
                    if (training_results):
                        ctrl_corrs[v] = corr(train_fmri[:, v], ctrl_preds[:, v])[0]
                        cs_corrs[v] = corr(train_fmri[:, v], cs_preds[:, v])[0]
                    else:
                        ctrl_corrs[v] = corr(test_fmri[:, v], ctrl_preds[:, v])[0]
                        cs_corrs[v] = corr(test_fmri[:, v], cs_preds[:, v])[0]

                # Save mean correlation
                mean_cs_corr = cs_corrs.mean()
                
                results_mat[model_subj_idx, target_subj_idx] = mean_cs_corr
                voxel_improvement_percentages_ctrl_mat[model_subj_idx, target_subj_idx] = np.count_nonzero(cs_corrs - ctrl_corrs > 0) / num_voxels_target_subj

                print(results_mat)
                print(voxel_improvement_percentages_ctrl_mat)
            
    if (save):
        if (training_results):
            accuracies_save_path = project_dir + "/results/" + roi + "_" + hemisphere_abbr + "h_cross_subject_cl_results_mat_training.npy"
            voxel_improvement_percentages_ctrl_save_path = project_dir + "/results/" + roi + hemisphere_abbr + "h_cs_voxel_improvement_percentages_cl_mat_training.npy"
        else:
            accuracies_save_path = project_dir + "/results/" + roi + hemisphere_abbr + "h_cross_subject_cl_results_mat.npy"
            voxel_improvement_percentages_ctrl_save_path = project_dir + "/results/" + roi + hemisphere_abbr + "h_cs_voxel_improvement_percentages_cl_mat.npy"
        np.save(accuracies_save_path, results_mat)
        np.save(voxel_improvement_percentages_ctrl_save_path, voxel_improvement_percentages_ctrl_mat)

        
    
# Get results for image classification task
def image_classification_results(project_dir, subj_num, hemisphere, rois, device, tuning_method='cl', dataset_name='caltech256', save=False):
    
    hemisphere_abbr = 'l' if hemisphere=='left' else 'r'
    
    save_path = project_dir + "/results/Subj" + str(subj_num) + "/subj" + str(subj_num) + "_" + hemisphere_abbr + "h_" + dataset_name + "_" + tuning_method + "_results.joblib"
    
    # Seed RNG, define image transforms for alexnet
    torch.manual_seed(0)
    alex_transform = transforms.Compose([
        transforms.Resize(256),
        transforms.CenterCrop(224),
        transforms.ToTensor(), # convert the images to a PyTorch tensor
        transforms.Normalize([0.485, 0.456, 0.406], [0.229, 0.224, 0.225]) # normalize the images color channels
                ])
    
    # Load the Data
    if (dataset_name=='caltech256'):
        image_dir = project_dir + "/caltech256"
        dataset = torchvision.datasets.ImageFolder(root=image_dir, transform=alex_transform)
    elif (dataset_name=='places365'):
        image_dir = project_dir + "/places365"
        dataset = torchvision.datasets.Places365(root=image_dir, split='val', small=True, transform=alex_transform)
    elif (dataset_name=='sun397'):
        image_dir = project_dir + "/sun397"
        dataset = torchvision.datasets.SUN397(root=image_dir, transform=alex_transform, download=True)
    elif (dataset_name=='imagenet'):
        image_dir = project_dir + "/imagenet"
        dataset = torchvision.datasets.ImageNet(root=image_dir, split='val', transform=alex_transform)
        
    total_num_images = len(dataset)
    shuffled_idxs = torch.randperm(total_num_images)
    
    train_size = int(0.85 * total_num_images)
    test_size = total_num_images - train_size
    
    train_idxs = shuffled_idxs[:train_size]
    test_idxs = shuffled_idxs[train_size:train_size+test_size]
    
    # Create train and test dataloaders
    train_dataset = torch.utils.data.Subset(dataset, train_idxs)
    test_dataset = torch.utils.data.Subset(dataset, test_idxs)
    train_dataloader = torch.utils.data.DataLoader(train_dataset, batch_size=256)
    test_dataloader = torch.utils.data.DataLoader(test_dataset, batch_size=256)
    
    
    # Get image features for untuned AlexNet
    from torchvision.models import AlexNet_Weights
    alex = torch.hub.load('pytorch/vision:v0.10.0', 'alexnet', weights=AlexNet_Weights.IMAGENET1K_V1)
    alex.to(device) 
    alex.eval() 
    
    feature_extractor = create_feature_extractor(alex, return_nodes=['classifier.5']).to(device)
    del alex
    
    # Get untuned alexnet features
    train_features_untuned = np.zeros((train_size, 4096))
    train_labels = np.zeros(train_size)
    for batch_index, data in tqdm(enumerate(train_dataloader), total=len(train_dataloader)):
        batch_size = data[0].shape[0]
        if (batch_index==0):
            low_idx = 0
            high_idx = batch_size
        else:
            low_idx = high_idx
            high_idx += batch_size
        # Extract features
        with torch.no_grad():
            ft = feature_extractor(data[0].to(device))
        # Flatten the features
        ft = torch.hstack([torch.flatten(l, start_dim=1) for l in ft.values()]).cpu().detach().numpy()
        train_features_untuned[low_idx:high_idx] = ft
        train_labels[low_idx:high_idx] = data[1]
        del ft
    
    test_features_untuned = np.zeros((test_size, 4096))
    test_labels = np.zeros(test_size)
    for batch_index, data in tqdm(enumerate(test_dataloader), total=len(test_dataloader)):
        batch_size = data[0].shape[0]
        if (batch_index==0):
            low_idx = 0
            high_idx = batch_size
        else:
            low_idx = high_idx
            high_idx += batch_size
        # Extract features
        with torch.no_grad():
            ft = feature_extractor(data[0].to(device))
        # Flatten the features
        ft = torch.hstack([torch.flatten(l, start_dim=1) for l in ft.values()]).cpu().detach().numpy()
        test_features_untuned[low_idx:high_idx] = ft
        del ft
        test_labels[low_idx:high_idx] = data[1]
        
    del feature_extractor
    
    scaler = StandardScaler()
    fit_scaler = scaler.fit(train_features_untuned)
    train_features_untuned = fit_scaler.transform(train_features_untuned)
    test_features_untuned = fit_scaler.transform(test_features_untuned)
    
    print("Fitting linear classifier...")
    classifier = LogisticRegression(max_iter=5000).fit(train_features_untuned, train_labels)
    preds = classifier.predict(test_features_untuned)
    acc = accuracy_score(test_labels, preds) * 100
    print("Untuned", acc)
    
    
    if (rois[0] == 'all'):
        rois = ["V1v", "V1d", "V2v", "V2d", "V3v", "V3d", "hV4", "EBA", "FBA-1", "FBA-2",
              "mTL-bodies", "OFA", "FFA-1", "FFA-2", "mTL-faces", "aTL-faces", "OPA",
         "PPA", "RSC", "OWFA", "VWFA-1", "VWFA-2", "mfs-words", "mTL-words",
       "early", "midventral", "midlateral", "midparietal", "ventral", "lateral", "parietal"]

    # Go through list of rois
    for roi in rois:
        
        # Get number of voxels for this model
        _, _, _, _, num_voxels = get_dataloaders(project_dir, device, subj_num, hemisphere, roi, 1024, shuffle=False)
        
        if (num_voxels < 20):
            print(roi, "is too small or empty")
        else:

            if (tuning_method=='cl'):
                model_dir = project_dir + r"/cl_models/Subj" + str(subj_num)
                model_path = model_dir + r"/subj" + str(subj_num) + "_" + hemisphere_abbr + "h_" + roi + "_model_e30.pt" 
                h_dim = int(num_voxels*0.8)
                z_dim = int(num_voxels*0.2)
                model = CLR_model(num_voxels, h_dim, z_dim)
                #model_path = project_dir + "/cl_models/Subj" + str(subj_num) + "/subj" + str(subj_num) + "_" + hemisphere_abbr + "h_" + roi + "_model_e30.pt"
                #model = torch.load(model_path).to(device)

            elif (tuning_method=='reg'):
                model_dir = project_dir + r"/baseline_models/nn_reg/Subj" + str(subj_num)
                model_path = model_dir + r"/subj" + str(subj_num) + "_" + hemisphere_abbr + "h_" + roi + "_reg_model_e75.pt" 
                model = fmri_reg(num_voxels)
           
            # Some models seem to be saved differently
            try:
                model.load_state_dict(torch.load(model_path)[0].state_dict())
            except:
                try:
                    model.load_state_dict(torch.load(model_path).state_dict())
                except:
                    model.load_state_dict(torch.load(model_path))

            model.to(device)
            model.eval()
            feature_extractor = tx.Extractor(model, ["alex.classifier.5"]).to(device)

            train_features_tuned = np.zeros((train_size, 4096))
            train_labels = np.zeros(train_size)
            for batch_index, data in tqdm(enumerate(train_dataloader), total=len(train_dataloader)):
                batch_size = data[0].shape[0]
                if (batch_index==0):
                    low_idx = 0
                    high_idx = batch_size
                else:
                    low_idx = high_idx
                    high_idx += batch_size
                # Extract features
                with torch.no_grad():
                    if (tuning_method=='cl'):
                        fmri_dummy = torch.zeros((batch_size, num_voxels)).to(device)
                        _, alex_out_dict = feature_extractor(fmri_dummy, data[0].to(device))
                    elif (tuning_method=='reg'):
                        _, alex_out_dict = feature_extractor(data[0].to(device))
                ft = alex_out_dict['alex.classifier.5'].detach().cpu().numpy()
                train_features_tuned[low_idx:high_idx] = ft
                train_labels[low_idx:high_idx] = data[1]
                del ft

            test_features_tuned = np.zeros((test_size, 4096))
            test_labels = np.zeros(test_size)
            for batch_index, data in tqdm(enumerate(test_dataloader), total=len(test_dataloader)):
                batch_size = data[0].shape[0]
                if (batch_index==0):
                    low_idx = 0
                    high_idx = batch_size
                else:
                    low_idx = high_idx
                    high_idx += batch_size
                # Extract features
                with torch.no_grad():
                    if (tuning_method=='cl'):
                        fmri_dummy = torch.zeros((batch_size, num_voxels)).to(device)
                        _, alex_out_dict = feature_extractor(fmri_dummy, data[0].to(device))
                    elif (tuning_method=='reg'):
                        _, alex_out_dict = feature_extractor(data[0].to(device))
                ft = alex_out_dict['alex.classifier.5'].detach().cpu().numpy()
                test_features_tuned[low_idx:high_idx] = ft
                test_labels[low_idx:high_idx] = data[1]
                del ft

            scaler = StandardScaler()
            fit_scaler = scaler.fit(train_features_tuned)
            train_features_tuned = fit_scaler.transform(train_features_tuned)
            test_features_tuned = fit_scaler.transform(test_features_tuned)

            print("Fitting linear classifier...")
            classifier = LogisticRegression(max_iter=5000).fit(train_features_tuned, train_labels)
            preds = classifier.predict(test_features_tuned)
            acc = accuracy_score(test_labels, preds) * 100
            #results[roi] = acc
            print(roi, acc)
            
            
            if (save):
                try:
                    # Load existing results, add result for roi if not already in the existing results
                    existing_results = joblib.load(save_path)
                    if roi not in existing_results.keys():
                        existing_results[roi] = acc
                        joblib.dump(existing_results, save_path)
                except:
                    results = {}
                    results[roi] = acc
                    joblib.dump(results, save_path)
                    
                    
# Choose best alpha for ridge regression via 5-fold cross validation (CL or reg)
def find_alpha(data_dir, device, subj_num, hemisphere, roi_name, model_type="cl"):
    print(roi_name) 
    alphas = list(np.logspace(-1, 6, num=8))
    results_dict = dict(zip(alphas, [0]*len(alphas)))
    total_acc_dict = dict(zip(alphas, [0]*len(alphas)))
    
    hemisphere_abbr = 'l' if hemisphere=='left' else 'r'
    
    for fold_num in range(5):
        
        #print("Fold " + str(fold_num+1) + ":")
        
        # Get train and val dataloaders for this fold number
        train_dataloader, val_dataloader, train_size, val_size, num_voxels = get_dataloaders_cv(data_dir, device, subj_num, hemisphere, roi_name, batch_size=1024, fold_num=fold_num)
        
        # Load model
        if (model_type == "cl"):
            model_dir = data_dir + r"/cl_models/Subj" + str(subj_num)
            model_path = model_dir + r"/subj" + str(subj_num) + "_" + hemisphere_abbr + "h_" + roi_name + "_model_e30.pt" 
            model = torch.load(model_path).to(device)
            # Create feature extractor from CL model
            feature_extractor = tx.Extractor(model, ["alex.classifier.6"]).to(device)
            del model
        elif (model_type == "reg"):
            model_dir = data_dir + r"/baseline_models/nn_reg/Subj" + str(subj_num)
            model_path = model_dir + r"/subj" + str(subj_num) + "_" + hemisphere_abbr + "h_" + roi_name + "_reg_model_e75.pt"
            # Create feature extractor from reg model
            feature_extractor = tx.Extractor(model, ["alex.classifier.6"]).to(device)
            del model
        elif (model_type == "untuned"):
            # Load best alexnet layer for control linear encoding model, create feature extractor
            best_alex_out_layer_path = data_dir + r"/best_alex_out_layers/Subj" + str(subj_num) + r"/subj" + str(subj_num) + "_" + hemisphere_abbr + "h_best_alex_layer_dict.joblib"
            best_alex_out_layer_dict = joblib.load(best_alex_out_layer_path)
            alex_out_layer_dims = {"features.2":46656, "features.5":32448, "features.7":64896, "features.9":43264,
                                   "features.12":9216, "classifier.2":4096, "classifier.5":4096, "classifier.6":1000}
            best_alex_layer = best_alex_out_layer_dict[roi_name]
            alex_out_size = alex_out_layer_dims[best_alex_layer]

            alex = torch.hub.load('pytorch/vision:v0.10.0', 'alexnet', weights=AlexNet_Weights.IMAGENET1K_V1)
            model.to(device) # send the model to the chosen device 
            model.eval() # set the model to evaluation mode
            feature_extractor = create_feature_extractor(model, return_nodes=[best_alex_layer])
            del model

            #Fit PCA using feature extractor
            pca = fit_pca(feature_extractor, train_dataloader, train_size, alex_out_size, batch_size=1024)

            # Get training and testing image pca features
            train_features = extract_pca_features(feature_extractor, pca, train_size, train_dataloader, 
                                                            is_img_dataloader=False, batch_size=1024)
            val_features = extract_pca_features(feature_extractor, pca, val_size, val_dataloader, 
                                                            is_img_dataloader=False, batch_size=1024)

    
        #print("Extracting image activations...")
        # Get training and tvalidation tuned alexnet features
        if (model_type != "untuned"):
            train_features = np.zeros((train_size, 1000))
            for batch_index, data in tqdm(enumerate(train_dataloader), total=len(train_dataloader)):
                batch_size = data[0].shape[0]
                if (batch_index==0):
                    low_idx = 0
                    high_idx = batch_size
                else:
                    low_idx = high_idx
                    high_idx += batch_size
                # Extract features
                with torch.no_grad():
                    if (model_type=='cl'):
                        _, alex_out_dict = feature_extractor(data[0], data[1])
                    elif (model_type=='reg'):
                        _, alex_out_dict = feature_extractor(data[1])
                ft = alex_out_dict['alex.classifier.6'].detach().cpu().numpy()
                train_features[low_idx:high_idx] = ft
                del ft

            val_features = np.zeros((val_size, 1000))
            for batch_index, data in tqdm(enumerate(val_dataloader), total=len(val_dataloader)):
                batch_size = data[0].shape[0]
                if (batch_index==0):
                    low_idx = 0
                    high_idx = batch_size
                else:
                    low_idx = high_idx
                    high_idx += batch_size
                # Extract features
                with torch.no_grad():
                    if (model_type=='cl'):
                        _, alex_out_dict = feature_extractor(data[0], data[1])
                    elif (model_type=='reg'):
                        _, alex_out_dict = feature_extractor(data[1])
                ft = alex_out_dict['alex.classifier.6'].detach().cpu().numpy()
                val_features[low_idx:high_idx] = ft
                del ft
        
        #print("Extracting fmri data...")
        train_fmri = get_fmri_from_dataloader(train_dataloader, train_size, num_voxels)
        val_fmri = get_fmri_from_dataloader(val_dataloader, val_size, num_voxels)    
        
        # Try different alpha values for ridge regression
        for alpha in alphas:

            #print("Fitting regression model...")
            encoding_model = Ridge(alpha=alpha).fit(train_features, train_fmri)
            preds = encoding_model.predict(val_features)
            
            #print("Getting validation fMRI data...")
            
            # Compute mean correlations for all methods
            corrs = np.zeros(num_voxels)
            #print("Computing correlations...")
            for v in tqdm(range(num_voxels)):
                corrs[v] = corr(val_fmri[:, v], preds[:, v])[0]

            avg_acc = corrs.mean()
            total_acc_dict[alpha] = total_acc_dict[alpha] + avg_acc
            results_dict[alpha] = total_acc_dict[alpha] / (fold_num + 1)

            #print(roi_name + ", alpha = " + str(alpha) + " val acc: " + str(avg_acc))
        
    print(roi_name, results_dict)



# Helper function to gather results into one list
def append_results_from_dict(existing_results, lh_results, rh_results):
    for key in (lh_results.keys()):
        if (lh_results[key] != -1):
            existing_results.append(lh_results[key])  
    for key in (rh_results.keys()):
        if (rh_results[key] != -1):
            existing_results.append(rh_results[key])  
    return existing_results

# Paired t test (sample size = 8 subjects for each ROI)
def t_test(project_dir):

    results_dir = project_dir + "/results"

    subj_nums = range(1,9)

    hemispheres = ["lh", "rh"]
    all_rois = ["V1v", "V1d", "V2v", "V2d", "V3v", "V3d", "hV4", "EBA", "FBA-1", "FBA-2",
            "mTL-bodies", "OFA", "FFA-1", "FFA-2", "mTL-faces", "aTL-faces", "OPA",
         "PPA", "RSC", "OWFA", "VWFA-1", "VWFA-2", "mfs-words", "mTL-words",
       "early", "midventral", "midlateral", "midparietal", "ventral", "lateral", "parietal"]
    
    # Collect results for each roi and method (key = roi, value = p-value)
    t_test_results_cl_vs_ctrl = {}
    t_test_results_cl_vs_reg = {}
    t_test_results_reg_vs_ctrl = {}

    for roi in all_rois:
        for hemisphere in hemispheres:
            ctrl_results = []
            cl_results = []
            reg_results = []
            for subj_num in subj_nums:
                ctrl_results.append(joblib.load(results_dir + "/Subj" + str(subj_num) + "/subj" + str(subj_num) + "_" + hemisphere + "_ctrl_results.joblib")[roi])
                cl_results.append(joblib.load(results_dir + "/Subj" + str(subj_num) + "/subj" + str(subj_num) + "_" + hemisphere + "_cl_results.joblib")[roi])
                reg_results.append(joblib.load(results_dir + "/Subj" + str(subj_num) + "/subj" + str(subj_num) + "_" + hemisphere + "_reg_results.joblib")[roi])
                
            ctrl_results = np.array(ctrl_results)
            cl_results = np.array(cl_results)
            reg_results = np.array(reg_results)

            roi_results_name = hemisphere + "_" + roi
            t_test_results_cl_vs_ctrl[roi_results_name] = ttest_rel(cl_results, ctrl_results)[1]
            t_test_results_cl_vs_reg[roi_results_name] = ttest_rel(cl_results, reg_results)[1]
            t_test_results_reg_vs_ctrl[roi_results_name] = ttest_rel(reg_results, ctrl_results)[1]
    
    t_test_results_cl_vs_ctrl_save_path = project_dir + "/results/t_test_results_cl_vs_ctrl_dict.joblib"
    t_test_results_cl_vs_reg_save_path = project_dir + "/results/t_test_results_cl_vs_reg_dict.joblib"
    t_test_results_reg_vs_ctrl_save_path = project_dir + "/results/t_test_results_reg_vs_ctrl_dict.joblib"

    joblib.dump(t_test_results_cl_vs_ctrl, t_test_results_cl_vs_ctrl_save_path)
    joblib.dump(t_test_results_cl_vs_reg, t_test_results_cl_vs_reg_save_path)
    joblib.dump(t_test_results_reg_vs_ctrl, t_test_results_reg_vs_ctrl_save_path)

# Function to create excel file with results in formatted table
# Function to create excel file with results in formatted table
def create_excel_results(project_dir):

    # First, create ROI-by-ROI results for each subject/hemisphere
    
    # Create new workbook
    wb = Workbook()

    subj_nums = [str(i) for i in range(1,9)]
    hemispheres = ["lh", "rh"]
    all_rois = ["V1v", "V1d", "V2v", "V2d", "V3v", "V3d", "hV4", "EBA", "FBA-1", "FBA-2",
            "mTL-bodies", "OFA", "FFA-1", "FFA-2", "mTL-faces", "aTL-faces", "OPA",
         "PPA", "RSC", "OWFA", "VWFA-1", "VWFA-2", "mfs-words", "mTL-words",
       "early", "midventral", "midlateral", "midparietal", "ventral", "lateral", "parietal"]
    first_data_row = 2
    last_data_row = first_data_row + len(all_rois) - 1
    avg_row = last_data_row + 1
    cols = ["A", "B", "C", "D", "E", "F", "G"]
    
    for subj_num in subj_nums:
        for hemisphere in hemispheres:
            wb.create_sheet("Subj" + subj_num + " " + hemisphere)
            ws = wb["Subj" + subj_num + " " + hemisphere]

            # Create titles
            titles = ["ROI", "Ctrl Avg.", "CL Avg.", "Reg Avg.", "% of voxels improved vs ctrl", "% of voxels improved vs reg"]
            for (col, title) in zip(cols, titles):
                ws[col + "1"] = title
                ws[col + "1"].font = Font(bold=True)
            for (idx, group) in enumerate(["All rois average", "Early rois average", "Higher rois average", "Anatomical rois average"]):
                ws["A" + str(avg_row + idx)] = group
                ws["A" + str(avg_row + idx)].font = Font(bold=True)

            # Load results
            ctrl_results = joblib.load(project_dir + "/results/Subj" + subj_num + "/subj" + subj_num + "_" + hemisphere + "_ctrl_results.joblib")
            cl_results = joblib.load(project_dir + "/results/Subj" + subj_num + "/subj" + subj_num + "_" + hemisphere + "_cl_results.joblib")
            reg_results = joblib.load(project_dir + "/results/Subj" + subj_num + "/subj" + subj_num + "_" + hemisphere + "_reg_results.joblib")
            cl_vs_ctrl_results = joblib.load(project_dir + "/results/Subj" + subj_num + "/subj" + subj_num + "_" + hemisphere + "_cl_vs_ctrl_improved_percentages.joblib") 
            cl_vs_reg_results = joblib.load(project_dir + "/results/Subj" + subj_num + "/subj" + subj_num + "_" + hemisphere + "_cl_vs_reg_improved_percentages.joblib") 

            # Insert results in sheet
            col_entries = [all_rois, ctrl_results, cl_results, reg_results, cl_vs_ctrl_results, cl_vs_reg_results]
            for (col, col_entry) in zip(cols, col_entries):
                # ROIs column
                if col == "A":
                    for (roi, row) in zip(all_rois, range(2, 2 + len(all_rois))):
                        ws[col + str(row)] = roi
                elif col in ["B", "C", "D"]:
                    for (roi, row) in zip(all_rois, range(2, 2 + len(all_rois))):
                        ws[col + str(row)] = col_entry[roi]

                        # Number formatting
                        if col_entry[roi] != -1:
                            ws[col + str(row)].number_format = '0.0000'
                        else:
                            ws[col + str(row)].number_format = '0'
                else:
                    for (roi, row) in zip(all_rois, range(2, 2 + len(all_rois))):
                        if col_entry[roi] != -1:
                            ws[col + str(row)] = col_entry[roi] * 100
                            ws[col + str(row)].number_format = '00.0'
                        else:
                            ws[col + str(row)] = -1
                            ws[col + str(row)].number_format = '0'

                # Fill out averages for roi groups (all, early, higher, and anatomical)
                if col != "A":
                    group_start_idxs = ["2", "2", "9", "26"]
                    group_end_idxs = ["32", "8", "25", "32", "32"]
                    for (idx, (group_start_idx, group_end_idx)) in enumerate(zip(group_start_idxs, group_end_idxs)):
                        ws[col + str(avg_row + idx)] = "=AVERAGEIF(" + col + group_start_idx + ":" + col + group_end_idx + ", \">=0\", " + col + group_start_idx + ":" + col + group_end_idx + ")"
                        if col in ["B", "C", "D"]:
                            ws[col + str(avg_row + idx)].number_format = '0.0000'
                        else:
                            ws[col + str(avg_row + idx)].number_format = '00.0'
                        ws[col + str(avg_row + idx)].font = Font(bold=True)
                        
    # Save file
    wb.save("Single_subject_results.xlsx")

    # Open file in data reading mode to get avg values for each subject

    # This part only works if excel is installed on the system
    excel_app = xlwings.App(visible=False)
    excel_book = excel_app.books.open("Single_subject_results.xlsx")
    excel_book.save()
    excel_book.close()
    excel_app.quit()


    wb = load_workbook("Single_subject_results.xlsx", data_only=True)
    for group in ["all", "early", "higher", "anatomical"]:
        make_avgs_excel_table(project_dir, wb, group)


# Group options are all, early, higher, anatomical
def make_avgs_excel_table(project_dir, workbook, group):

    subj_nums = [str(i) for i in range(1,9)]
    hemispheres = ["lh", "rh"]
    cols = ["A", "B", "C", "D", "E", "F"]

    # Collect averages for each subject in lists
    subj_avgs = {"Avg. Ctrl. Acc.":[], "Avg. CL Acc.":[], "Avg. Reg Acc.":[], "Avg. % of voxels improved vs ctrl":[], "Avg. % of voxels improved vs reg":[]}
    cols_to_avg = {"B":"Avg. Ctrl. Acc.", "C":"Avg. CL Acc.", "D":"Avg. Reg Acc.", "E":"Avg. % of voxels improved vs ctrl", "F":"Avg. % of voxels improved vs reg"}
    group_avg_rows = {"all":"33", "early":"34", "higher":"35", "anatomical":"36"}
    row = group_avg_rows[group]
    for subj_num in subj_nums:
        lh_values = {"Avg. Ctrl. Acc.":[], "Avg. CL Acc.":[], "Avg. Reg Acc.":[], "Avg. % of voxels improved vs ctrl":[], "Avg. % of voxels improved vs reg":[]}
        for hemisphere in hemispheres:
            ws = workbook["Subj" + subj_num + " " + hemisphere]
            for col in cols[1:]:
                if (hemisphere == "lh"):
                    lh_values[cols_to_avg[col]] = float(ws[col + row].value)
                else:
                    subj_avgs[cols_to_avg[col]].append((lh_values[cols_to_avg[col]] + float(ws[col + row].value)) / 2)
                    
    
    title = "Subj avgs " + group
    workbook.create_sheet(title)
    ws = workbook[title]
    titles = ["Subject", "Avg. Ctrl. Acc.", "Avg. CL Acc.", "Avg. Reg Acc.", "Avg. % of voxels improved vs ctrl", "Avg. % of voxels improved vs reg"]
    # Add titles
    for (col, title) in zip(cols, titles):
        ws[col + "1"] = title
        ws[col + "1"].font = Font(bold=True)
    # Fill in data
    for subj_num in subj_nums:
        for (col, title) in zip(cols, titles):
            row = str(int(subj_num) + 1)
            if col == "A":
                ws[col + row] = int(subj_num)
            elif col in ["B", "C", "D"]:
                ws[col + row] = subj_avgs[title][int(subj_num) - 1]
                ws[col + row].number_format = '0.000'
            elif col in ["E", "F"]:
                ws[col + row] = subj_avgs[title][int(subj_num) - 1]
                ws[col + row].number_format = '00.0'

    # Final averages
    for col in cols:
        if col == "A":
            ws[col + "10"] = "All"
        elif (col in ["B", "C", "D"]):
            ws[col + "10"] = "=AVERAGE(" + col + "2:" + col + "9" + ")"
            ws[col + "10"].number_format = '0.000'
        elif (col in ["E", "F"]):
            ws[col + "10"] = "=AVERAGE(" + col + "2:" + col + "9" + ")"
            ws[col + "10"].number_format = '00.0'
        ws[col + "10"].font = Font(bold=True)

    workbook.save("Single_subject_results.xlsx")



# Function to create heatmaps for cross-subject results
#def cs_heatmap(project_dir, data, roi_name, method, improvement_percentages=True, save=False, average=False):
def cs_heatmap(project_dir, roi, hemisphere):

    hemisphere_abbr = 'lh' if hemisphere == 'left' else 'rh'
    roi_name = hemisphere_abbr.upper() + " " + roi

    try:
        file_path = project_dir + "/results/" + roi + "_" + hemisphere_abbr + "_cs_voxel_improvement_percentages_cl_mat.npy"
        cs_results = np.load(file_path)
    except:
        file_path = project_dir + "/results/" + roi + hemisphere_abbr + "_cs_voxel_improvement_percentages_cl_mat.npy"
        cs_results = np.load(file_path)

    num_subjs = 8
    
    # Meet figure requirements for neural computation manuscript (min. 600 dpi, figsize = 4.25x7 in)
    fig, axs = plt.subplots(1, num_subjs, figsize=(4.25,4.25), gridspec_kw={'wspace': 0}, dpi=600)

    subj_labels = [str(x) for x in range(1, num_subjs+1)]
    
    cmap = 'Reds'
    cmap_grad_tol = 0.01
    
    counter = 0
    sns.set(font_scale=0.6)
    for column in range(num_subjs):
        col_data = cs_results[:, column]
        if (column==0):
            sns.heatmap(col_data.reshape(num_subjs,1), yticklabels=subj_labels, xticklabels=[subj_labels[counter]], ax=axs[counter],
                    annot=True, fmt='.3f', cmap=cmap, cbar=False, vmin=np.min(col_data)-cmap_grad_tol, vmax=np.max(col_data)+cmap_grad_tol)
        else:
            sns.heatmap(col_data.reshape(num_subjs,1), yticklabels=[], xticklabels=[subj_labels[counter]], ax=axs[counter],
                    annot=True, fmt='.3f', cmap=cmap, cbar=False, vmin=np.min(col_data)-cmap_grad_tol, vmax=np.max(col_data)+cmap_grad_tol)
        counter += 1


    fig.supylabel("Model Subject")
    fig.supxlabel("Evaluation Subject")
    #if (improvement_percentages):
    title = "Cross Subject Voxel Improvement Percentages vs. Ctrl for " + roi_name 
    #else:
    #    title = "Cross Subject Average Correlations for " + roi_name + " (" + method + ")"
    plt.suptitle(title)
    
    results_dir = project_dir + "/results/"
    fig_save_path = results_dir + roi + "_" + hemisphere_abbr + "_cs_picture_cl"
    plt.savefig(fig_save_path)
    

# Plot average cross-subject voxel improvement percentages across all ROIs, split by group (early, higher, anatomical, all)
def cs_heatmap_avgs(project_dir, group='all'):

    hemisphere_abbrs = ['lh', 'rh']
    num_subjs = 8

    #all_rois = ["V1v", "V1d", "V2v", "V2d", "V3v", "V3d", "hV4", "EBA", "FBA-1", "FBA-2",
    #        "mTL-bodies", "OFA", "FFA-1", "FFA-2", "mTL-faces", "aTL-faces", "OPA",
     #           "PPA", "RSC", "OWFA", "VWFA-1", "VWFA-2", "mfs-words", "mTL-words",
     #       "early", "midventral", "midlateral", "midparietal", "ventral", "lateral", "parietal"]

     
    # Only include ROIs which are present in all subjects
    if group == 'early':
        rois = ["V1v", "V1d", "V2v", "V2d", "V3v", "V3d", "hV4"]
    elif group == 'higher':
        rois = ["EBA", "FBA-1", "FBA-2", "mTL-bodies", "OFA", "FFA-1", "FFA-2", "mTL-faces", "aTL-faces", "OPA",
                "PPA", "RSC", "OWFA", "VWFA-1", "VWFA-2", "mfs-words", "mTL-words"]
    elif group == 'anatomical':
        rois = ["early", "midventral", "midlateral", "midparietal", "ventral", "lateral", "parietal"]
    elif group == 'all':
        rois = ["V1v", "V1d", "V2v", "V2d", "V3v", "V3d", "hV4", "EBA", "FBA-1", "FBA-2",
            "mTL-bodies", "OFA", "FFA-1", "FFA-2", "mTL-faces", "aTL-faces", "OPA",
                "PPA", "RSC", "OWFA", "VWFA-1", "VWFA-2", "mfs-words", "mTL-words",
            "early", "midventral", "midlateral", "midparietal", "ventral", "lateral", "parietal"]
    else:
        print("Invalid group!")
        return


    roi_counter = 0
    roi_results = []
    for roi in rois:
        for hemisphere_abbr in hemisphere_abbrs:
            try:
                file_path = project_dir + "/results/" + roi + "_" + hemisphere_abbr + "_cs_voxel_improvement_percentages_cl_mat.npy"
                roi_results.append(np.load(file_path))
                roi_counter += 1
            except:
                try:
                    file_path = project_dir + "/results/" + roi + hemisphere_abbr + "_cs_voxel_improvement_percentages_cl_mat.npy"
                    roi_results.append(np.load(file_path))
                    roi_counter += 1
                except:
                    pass
            
            
    all_results = np.zeros((num_subjs, num_subjs, roi_counter))
    print("Number of rois included for " + group + " group: "  + str(roi_counter))
    for counter, roi_result in enumerate(roi_results):
        all_results[:,:, counter] = roi_result

    result_avgs = np.mean(all_results, axis=2)

    # Meet figure requirements for neural computation manuscript (min. 600 dpi, figsize = 4.25x7 in)
    fig, axs = plt.subplots(1, num_subjs, figsize=(4.25,4.25), gridspec_kw={'wspace': 0}, dpi=600)

    subj_labels = [str(x) for x in range(1, num_subjs+1)]
    
    cmap = 'Reds'
    cmap_grad_tol = 0.01
    
    counter = 0
    sns.set(font_scale=0.6)
    for column in range(num_subjs):
        col_data = result_avgs[:, column]
        if (column==0):
            sns.heatmap(col_data.reshape(num_subjs,1), yticklabels=subj_labels, xticklabels=[subj_labels[counter]], ax=axs[counter],
                    annot=True, fmt='.3f', cmap=cmap, cbar=False, vmin=np.min(col_data)-cmap_grad_tol, vmax=np.max(col_data)+cmap_grad_tol)
        else:
            sns.heatmap(col_data.reshape(num_subjs,1), yticklabels=[], xticklabels=[subj_labels[counter]], ax=axs[counter],
                    annot=True, fmt='.3f', cmap=cmap, cbar=False, vmin=np.min(col_data)-cmap_grad_tol, vmax=np.max(col_data)+cmap_grad_tol)
        counter += 1


    fig.supylabel("Model Subject")
    fig.supxlabel("Evaluation Subject")
    if (group != 'all'):
        title = "Avg CS Voxel Improvement Percentages vs. Ctrl for " + group + " group"
    else:
        title = "Avg CS Voxel Improvement Percentages vs. Ctrl for all ROIs"
    plt.suptitle(title)
    
    results_dir = project_dir + "/results/"
    fig_save_path = results_dir + group + "_group_cs_avgs_picture"
    plt.savefig(fig_save_path)
                
